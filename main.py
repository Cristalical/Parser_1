from bs4 import BeautifulSoup
import requests
from openpyxl import Workbook
import logging

from aiogram import Bot, Dispatcher, F
from aiogram.enums.parse_mode import ParseMode
from aiogram.client.default import DefaultBotProperties
from aiogram.fsm.storage.memory import MemoryStorage
from aiogram.filters import Command
from aiogram.types import Message, FSInputFile, ReplyKeyboardMarkup, KeyboardButton, InlineKeyboardButton, InlineKeyboardMarkup, CallbackQuery
from aiogram.fsm.state import State, StatesGroup
from aiogram.fsm.context import FSMContext
import asyncio

# Настройка логирования
logging.basicConfig(level=logging.INFO)

# Инициализация бота
bot = Bot(token="8124940732:AAHq9s8__mKQ5egnH1R7S7ztEKriK2g0F1U", default=DefaultBotProperties(parse_mode=ParseMode.HTML))
dp = Dispatcher(storage=MemoryStorage())

# Обработчик команды /start
@dp.message(Command("start"))
async def cmd_start(message: Message):
    keyboard = ReplyKeyboardMarkup(
        keyboard=[
            [KeyboardButton(text="Сделать запрос")],
            [KeyboardButton(text="Посмотреть MCC")]
        ],
        resize_keyboard=True
    )
    await message.answer('''Привет!\nЭтот бот создан для быстрого получения MCC кодов.\nЧтобы перейти к поиску, нажми кнопку или напиши /s.\nДля просмотра существующих MCC напиши /p или воспользуйся кнопкой.''', reply_markup=keyboard)

@dp.message(F.text == "Сделать запрос")
async def command_s(message: Message, state: FSMContext):
    await cmd_s(message, state)

@dp.message(F.text == "Посмотреть MCC")
async def command_p(message: Message, state: FSMContext):
    await cmd_p(message, state)

# Обработчик команды /p
@dp.message(Command("p"))
async def cmd_p(message: Message, state: FSMContext):
    keyboard = InlineKeyboardMarkup(
        inline_keyboard=[
        [InlineKeyboardButton(text="Транспорт", callback_data='transport')],
        [InlineKeyboardButton(text="Фаст Фуд", callback_data='fast_food')],
        [InlineKeyboardButton(text="Фото, Видео", callback_data='photo_video')],
        [InlineKeyboardButton(text="Цветы", callback_data='flowers')],
        [InlineKeyboardButton(text="Duty Free", callback_data='duty_free')],
    ],
        resize_keyboard=True
    )
    await message.answer('Выберите категорию', reply_markup=keyboard)

@dp.callback_query(F.data == 'transport')
async def subjects(call: CallbackQuery):
    await call.answer()
    await call.message.answer('''4111, 4121, 4131, 4457, 4468, 4784, 4789, 5013, 5271, 5551, 5561, 5592, 5598, 5599, 7511, 7523''')

@dp.callback_query(F.data == 'fast_food')
async def subjects(call: CallbackQuery):
    await call.answer()
    await call.message.answer('''5814''')

@dp.callback_query(F.data == 'photo_video')
async def subjects(call: CallbackQuery):
    await call.answer()
    await call.message.answer('''5044, 5045, 5946, 7332, 7333, 7338, 7339, 7395''')

@dp.callback_query(F.data == 'flowers')
async def subjects(call: CallbackQuery):
    await call.answer()
    await call.message.answer('''5193, 5992''')

@dp.callback_query(F.data == 'duty_free')
async def subjects(call: CallbackQuery):
    await call.answer()
    await call.message.answer('''5309''')

# Состояние для FSM
class Form(StatesGroup):
    waiting_for_input = State()

# Обработчик команды /s
@dp.message(Command("s"))
async def cmd_s(message: Message, state: FSMContext):
    await message.answer('Введите ваш запрос:')
    await state.set_state(Form.waiting_for_input)

# Обработчик ввода пользователя
@dp.message(Form.waiting_for_input)
async def process(message: Message, state: FSMContext):
    user_input = message.text
    await state.clear()
    url_1 = "https://mcc-codes.ru/search/?q="
    req = user_input
    url_1 += req
    response_1 = requests.get(url_1)
    soup_1 = BeautifulSoup(response_1.text, 'lxml')
    rows_1 = soup_1.find_all('tr')
    data_list = []
    for row in rows_1:
        if 'td' in str(row):
            cells = row.find_all('td')
            row_data = [cell.get_text(strip=True) for cell in cells]
            data_list.append(row_data)
    save(data_list)
    doc = FSInputFile("result_mcc.xlsx")
    await message.reply_document(doc)

# Функция для сохранения данных в Excel
def save(lst):
    # Имя файла
    filename = 'result_mcc.xlsx'

    # Создание нового Workbook
    wb = Workbook()
    sheet = wb.active

    sheet.column_dimensions['A'].width = 5
    sheet.column_dimensions['B'].width = 25
    sheet.column_dimensions['C'].width = 25
    sheet.column_dimensions['D'].width = 12

    # Вводим таблицы
    c = sheet.cell(row=1, column=1)
    c1 = sheet.cell(row=1, column=2)
    c2 = sheet.cell(row=1, column=3)
    c3 = sheet.cell(row=1, column=4)

    # Ввод названий столбцов
    c.value = "MCC"
    c1.value = "Название точки"
    c2.value = "Адрес оплаты"
    c3.value = "Актуалы"

    i = 2  # Начинаем с второй строки, так как первая строка - заголовки
    for row_data in lst:
        c = sheet.cell(row=i, column=1)
        c1 = sheet.cell(row=i, column=2)
        c2 = sheet.cell(row=i, column=3)
        c3 = sheet.cell(row=i, column=4)

        c.value = int(row_data[0])
        c1.value = row_data[1]
        c2.value = row_data[2]
        c3.value = row_data[3][:10]
        i += 1

    # Сохранение данных в таблицу Excel
    try:
        wb.save(filename)
        logging.info(f"Файл {filename} успешно сохранен.")
    except Exception as e:
        logging.error(f"Ошибка при сохранении файла {filename}: {e}")
        raise

# Запуск бота
async def main():
    await dp.start_polling(bot)

if __name__ == '__main__':
    asyncio.run(main())