from openpyxl import Workbook
import logging

# Функция для сохранения данных в Excel
def save(lst):
    # Имя файла
    filename = 'result_mcc.xlsx'

    # Создание нового Workbook
    wb = Workbook()
    sheet = wb.active

    sheet.column_dimensions['A'].width = 5
    sheet.column_dimensions['B'].width = 25
    sheet.column_dimensions['C'].width = 25
    sheet.column_dimensions['D'].width = 12

    # Вводим таблицы
    c = sheet.cell(row=1, column=1)
    c1 = sheet.cell(row=1, column=2)
    c2 = sheet.cell(row=1, column=3)
    c3 = sheet.cell(row=1, column=4)

    # Ввод названий столбцов
    c.value = "MCC"
    c1.value = "Название точки"
    c2.value = "Адрес оплаты"
    c3.value = "Актуалы/ссылки"

    i = 2  # Начинаем с второй строки, так как первая строка - заголовки
    for row_data in lst:
        c = sheet.cell(row=i, column=1)
        c1 = sheet.cell(row=i, column=2)
        c2 = sheet.cell(row=i, column=3)
        c3 = sheet.cell(row=i, column=4)

        c.value = int(row_data[0])
        c1.value = row_data[1]
        c2.value = row_data[2]
        c3.value = row_data[3][:10] if row_data[3][:7] != 'http://' else row_data[3]
        i += 1

    # Сохранение данных в таблицу Excel
    try:
        wb.save(filename)
        logging.info(f"Файл {filename} успешно сохранен.")
    except Exception as e:
        logging.error(f"Ошибка при сохранении файла {filename}: {e}")
        raise